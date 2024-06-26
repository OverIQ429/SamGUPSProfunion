from django.utils import timezone
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin
from django.contrib.auth.models import User, Group
from django.contrib.auth.views import LogoutView
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.urls import reverse, reverse_lazy
from django.views import View
from django.views.generic import TemplateView, CreateView, UpdateView, ListView, DetailView
from openpyxl import Workbook
from .models import Profile, Appends, Decision, News, Photo
from .forms import ProfileEditForm, CustomUserCreationForm, AppendsEditForm, NewForm
from io import BytesIO
from django.db.models import Q

def login_view(request: HttpRequest):
    if request.method == "GET":
        if request.user.is_authenticated:
            if request.user.is_superuser:
                return redirect('/admin/')
            else:
                return redirect('/about-me/')

        return render(request, 'myauth/login.html')

    username = request.POST['username']
    password = request.POST['password']

    user = authenticate(request, username=username, password=password)
    if user is not None:
        login(request, user)
        if user.is_superuser:
            return redirect("/admin/")
        else:
            return redirect('/about-me/')

    return render(request, 'myauth/login.html', {"error": "Invalid login credentials"})

def logout_view(request: HttpRequest):
    logout(request)
    return redirect(reverse("myauth:login"))

class MyLogoutView(LogoutView):
    next_page = reverse_lazy("myauth:login")

class AboutMeView(TemplateView):
    template_name = "myauth/about-me.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['appends'] = Appends.objects.filter(user=self.request.user)
        return context

class RegisterView(CreateView):
    form_class = CustomUserCreationForm
    template_name = 'myauth/register.html'
    success_url = reverse_lazy("myauth:about-me")

    def form_valid(self, form):
        response = super().form_valid(form)
        profile = Profile.objects.create(user=self.object)
        profile.email = form.cleaned_data.get("email")
        profile.phone_number = form.cleaned_data.get("phone_number")
        profile.name = form.cleaned_data.get("name")
        profile.secondname = form.cleaned_data.get("secondname")
        profile.lastname = form.cleaned_data.get("lastname")
        profile.save()
        username = form.cleaned_data.get("username")
        password = form.cleaned_data.get("password1")
        user = authenticate(
            self.request, username=username, password=password
        )
        login(request=self.request, user=user)
        group, created = Group.objects.get_or_create(name='Not union group')
        user.groups.add(group)

        return response

class AvatarUpdateView(UserPassesTestMixin, UpdateView):
    def test_func(self):
        if (self.request.user.id == self.get_object().user.profile.id
                or self.request.user.is_staff):
            return True

    model = Profile
    template_name = "myauth/avatar_update.html"
    form_class = ProfileEditForm

    def get_success_url(self):
        if self.request.user.is_staff:
            return reverse("myauth:users_list",)
        else:
            return reverse("myauth:about-me", )
class UsersList(UserPassesTestMixin, ListView):

    def test_func(self):
        if self.request.user.is_staff:
            return True

    template_name = "myauth/users_list.html"
    context_object_name = "profiles"
    queryset = User.objects.all()

    def get_queryset(self):
        queryset = super().get_queryset().exclude(is_superuser=True)

        name = self.request.GET.get('name')

        if name:
            queryset = queryset.filter(Q(profile__name__icontains=name) |
                                       Q(profile__user__lastname__icontains=name) |
                                       Q(profile__user__second__icontains=name)|
                                       Q(profile__user__email__icontains=name)|
                                       Q(profile__user__group__icontains=name)|
                                       Q(profile__user__group__faculty__icontains=name)|
                                       Q(profile__user__group__course__icontains=name))
        return queryset

class UserDetail(UserPassesTestMixin, DetailView):
    def test_func(self):
        if self.request.user.is_staff:
            return True
    model = User
    template_name = "myauth/users_detail.html"
    context_object_name = "user"

    def get_context_data(self, **kwargs):
        context = super(UserDetail, self).get_context_data(**kwargs)
        context['appends'] = Appends.objects.filter(user=self.get_object())
        return context
class Create_Append(LoginRequiredMixin, CreateView):
    model = Appends
    fields = "description", "file"
    success_url = reverse_lazy("myauth:about-me")
    def form_valid(self, form):
        decision, created = Decision.objects.get_or_create(text="Не рассмотрено")
        form.instance.decision = decision
        form.instance.user = self.request.user
        return super(Create_Append, self).form_valid(form)

class All_Appends(TemplateView):
    template_name = "myauth/appends.html"

class CheckAppens(UserPassesTestMixin,UpdateView):
    def test_func(self):
        if self.request.user.is_staff:
            return True

    model = Appends
    template_name = "myauth/append_update.html"
    form_class = AppendsEditForm
    def get_success_url(self):
        append = self.object
        return reverse('myauth:user_detail', kwargs={'pk': append.user.pk})

    def form_valid(self, form):
        append = form.save(commit=False)
        decision_text = self.request.POST.get('decision')
        if decision_text:
            if append.decision.text == "Одобренно":
                append.user.profile.is_union_member = True
                append.user.profile.save()
                group, created = Group.objects.get_or_create(name='Union Members')
                append.user.groups.add(group)
                previous_group = Group.objects.get(name='Not union group')
                append.user.groups.remove(previous_group)
        append.save()
        return super().form_valid(form)
class Check_All_Appends(UserPassesTestMixin, ListView):
    def test_func(self):
        if self.request.user.is_staff:
            return True

    template_name = "myauth/allappends_list.html"
    context_object_name = "profiles"

    def get_queryset(self):
        return Profile.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['appends'] = Appends.objects.select_related("user",'description', 'decision').all()
        return context

class MyAppends(LoginRequiredMixin, ListView):
    model = Appends
    template_name = "myauth/my_appends.html"

    def get_queryset(self):
        return Appends.objects.filter(user=self.request.user)

def delete_append(request, append_id):
    append = get_object_or_404(Appends, pk=append_id)
    append.delete()
    return redirect('myauth:my_appends')

class ReportofProfiles(UserPassesTestMixin, ListView):
    def test_func(self):
        if self.request.user.is_staff:
            return True

    template_name = "myauth/report_list.html"
    context_object_name = "profiles"
    queryset = User.objects.all()

    def get_queryset(self):
        queryset = super().get_queryset().exclude(is_superuser=True)

        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        course = self.request.GET.get('course')
        faculty = self.request.GET.get('faculty')
        group_id = self.request.GET.get('group')
        username = self.request.GET.get('username')

        if start_date:
            queryset = queryset.filter(
                profile__created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(
                profile__created_at__lte=end_date)
        if course:
            queryset = queryset.filter(profile__group__course__icontains=course)
        if faculty:
            queryset = queryset.filter(profile__group__faculty__icontains=faculty)
        if group_id:
            queryset = queryset.filter(profile__group_id=group_id)
        if username:
            queryset = queryset.filter(Q(profile__user__username__icontains=username) |
                                       Q(profile__user__first_name__icontains=username) |
                                       Q(profile__user__last_name__icontains=username))
        return queryset


    def get(self, request, *args, **kwargs):
        if 'excel' in request.GET:
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="profiles.xlsx"'
            wb = Workbook()
            ws = wb.active
            ws.append(['Имя', 'Фамилия', 'Отчество', 'Почта', 'Группа', 'Факультет', 'Курс', 'Дата регистрации', 'Член профсоюза'])

            for user in self.get_queryset():
                if not user.is_superuser:
                    if hasattr(user, 'profile') and user.profile is not None:
                        group = str(getattr(user.profile, 'group', ''))
                        faculty = str(getattr(user.profile.group, 'faculty', ''))
                        course = str(getattr(user.profile.group, 'course', ''))
                        is_union_member = str(getattr(user.profile, 'is_union_member', ''))
                        created = str(getattr(user.profile, 'created_at', ''))
                        ws.append([user.username, user.first_name, user.last_name, user.email, group, faculty, course, created, is_union_member])
                    else:
                        ws.append([user.username, user.first_name, user.last_name, user.email, "", "", "", "",""])
            wb.save(response)
            return response

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'start_date' in self.request.GET:
            context['start_date'] = self.request.GET['start_date']
        if 'end_date' in self.request.GET:
            context['end_date'] = self.request.GET['end_date']
        if 'group' in self.request.GET:
            context['group'] = self.request.GET['group']
        return context


class DownloadTableView(View):
    def get(self, request, *args, **kwargs):
        queryset = User.objects.all()

        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        group_id = request.GET.get('group')

        if start_date:
            queryset = queryset.filter(profile__created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(profile__created_at__lte=end_date)
        if group_id:
            queryset = queryset.filter(profile__group_id=group_id)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="filtered_table.csv"'

        import csv
        writer = csv.writer(response)

        wb = Workbook()
        ws = wb.active
        ws.append(['Имя', 'Фамилия', 'Отчество', 'Email', 'Группа', 'Факультет', 'Курс', 'Дата регистрации', 'Член профсоюза'])

        for user in queryset:
            if not user.is_superuser:
                group = str(getattr(user.profile, 'group', ''))
                faculty = str(getattr(user.profile.group, 'faculty', ''))
                course = str(getattr(user.profile.group, 'course', ''))
                is_union_member = str(getattr(user.profile, 'is_union_member', ''))
                created = str(getattr(user.profile, 'created_at', ''))
                ws.append([user.username, user.first_name, user.last_name, user.email, group, faculty, course, created, is_union_member])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="filtered_table.xlsx"'

        return response

class ReportofAppends(UserPassesTestMixin, ListView):
    def test_func(self):
        if self.request.user.is_staff:
            return True

    template_name = "myauth/order_report_list.html"
    context_object_name = "appends"
    queryset = Appends.objects.all()

    def get_queryset(self):
        queryset = Appends.objects.all()

        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        description = self.request.GET.get('description')
        username = self.request.GET.get('username')
        decision = self.request.GET.get('decision')

        if start_date:
            queryset = queryset.filter(
                created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(
                created_at__lte=end_date)
        if description:
            queryset = queryset.filter(description__text__icontains=description)
        if decision:
            queryset = queryset.filter(decision__text__icontains=decision)
        if username:
            queryset = queryset.filter(Q(user__username__icontains=username))
        return queryset


    def get(self, request, *args, **kwargs):
        if 'excel' in request.GET:
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="profiles.xlsx"'
            wb = Workbook()
            ws = wb.active
            ws.append(['Имя Пользователя', 'Назавние', 'Решение', 'Дата подачи'])

            for append in self.get_queryset():
                username = append.user.username
                description = append.description.text if append.description else ''
                decision = append.decision.text if append.decision else ''
                created_at = append.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ws.append([username, description, decision, created_at])

            wb.save(response)
            return response

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'start_date' in self.request.GET:
            context['start_date'] = self.request.GET['start_date']
        if 'end_date' in self.request.GET:
            context['end_date'] = self.request.GET['end_date']
        if 'group' in self.request.GET:
            context['group'] = self.request.GET['group']
        return context


class DownloadTableAppendView(View):
    def get(self, request, *args, **kwargs):
        queryset = Appends.objects.all()

        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        if start_date:
            queryset = queryset.filter(profile__created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(profile__created_at__lte=end_date)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="appends_table.csv"'

        wb = Workbook()
        ws = wb.active
        ws.append(['Имя пользователя', 'Описание', 'Файл', 'Решение', 'Комментарий', 'Дата создания'])

        for append in queryset:
            username = append.user.username
            description = append.description.text if append.description else ''  # Предполагаем, что у Description есть поле text
            file_url = request.build_absolute_uri(append.file.url) if append.file else ''
            decision = append.decision.text if append.decision else ''  # Предполагаем, что у Decision есть поле text
            commentary = append.commentary
            created_at = append.created_at.strftime('%Y-%m-%d %H:%M:%S')

            ws.append([username, description, file_url, decision, commentary, created_at])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="appends_table.xlsx"'

        return response

class CreateNew(LoginRequiredMixin, CreateView):
    model = News
    fields = ["type", "name", "article", "avatar", "document"]
    template_name = "myauth/create_news.html"
    success_url = reverse_lazy("myauth:about-me")

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        photo_files = request.FILES.getlist('photos')  # Получаем список загруженных файлов фотографий
        if form.is_valid():
            self.object = form.save()
            for file in photo_files:
                photo = Photo(image=file)
                photo.save()
                self.object.photos.add(photo)
            return self.form_valid(form)
        else:
            return self.form_invalid(form)
